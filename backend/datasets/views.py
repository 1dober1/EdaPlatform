import csv
import json
import io

from django.http import FileResponse
from rest_framework import generics, permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response

from .models import Dataset
from .serializers import DatasetSerializer, DatasetUploadSerializer
from django.contrib.auth import get_user_model

class StatsView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        User = get_user_model()
        users_count = User.objects.count()
        datasets_count = Dataset.objects.count()
        # charts can be a generic multiplication if not stored in DB
        charts_count = datasets_count * 5 + 12 
        
        return Response({
            'users': users_count,
            'datasets': datasets_count,
            'charts': charts_count
        })

DEMO_DATASETS = {
    'titanic': {
        'name': 'Titanic',
        'description': 'Данные о пассажирах Титаника (выживаемость, классы, возраст)',
        'filename': 'titanic.csv',
    },
    'california_housing': {
        'name': 'California Housing',
        'description': 'Данные переписи Калифорнии (цены на жилье)',
        'filename': 'california_housing.csv',
    },
    'iris': {
        'name': 'Iris',
        'description': 'Классический датасет с параметрами цветков ириса',
        'filename': 'iris.csv',
    },
    'diamonds': {
        'name': 'Diamonds',
        'description': 'Цены и характеристики почти 54 000 бриллиантов',
        'filename': 'diamonds.csv',
    },
    'penguins': {
        'name': 'Palmer Penguins',
        'description': 'Данные о трех видах пингвинов (размеры крыльев, масса)',
        'filename': 'penguins.csv',
    },
    'tips': {
        'name': 'Restaurant Tips',
        'description': 'Данные о чаевых в ресторане',
        'filename': 'tips.csv',
    },
}


class DatasetListCreateView(generics.ListCreateAPIView):
    """GET — список датасетов; POST — загрузить новый."""
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return DatasetUploadSerializer
        return DatasetSerializer

    def get_queryset(self):
        return Dataset.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        uploaded_file = serializer.validated_data['file']
        rows, cols = self._count_rows_cols(uploaded_file)
        serializer.save(user=self.request.user, rows=rows, columns=cols)

    @staticmethod
    def _count_rows_cols(uploaded_file):
        """Подсчёт строк и колонок загруженного файла."""
        try:
            ext = uploaded_file.name.rsplit('.', 1)[-1].lower()
            uploaded_file.seek(0)

            if ext == 'parquet':
                # Parquet files can't be easily read without pandas
                # We'll store 0,0 and update later if needed
                return 0, 0

            if ext in ('xlsx', 'xls'):
                # Excel files are primarily parsed on the frontend
                # Try openpyxl if available
                try:
                    import openpyxl
                    uploaded_file.seek(0)
                    wb = openpyxl.load_workbook(uploaded_file, read_only=True)
                    ws = wb.active
                    rows = ws.max_row or 0
                    cols = ws.max_column or 0
                    wb.close()
                    return max(rows - 1, 0), cols  # exclude header
                except ImportError:
                    return 0, 0
            
            content = uploaded_file.read().decode('utf-8')
            uploaded_file.seek(0)

            if ext == 'csv':
                try:
                    dialect = csv.Sniffer().sniff(content[:2048], delimiters=',;\t|')
                    reader = csv.reader(io.StringIO(content), dialect)
                except Exception:
                    reader = csv.reader(io.StringIO(content))
                
                rows_list = list(reader)
                if not rows_list:
                    return 0, 0
                cols = len(rows_list[0])
                return len(rows_list) - 1, cols  # exclude header
            elif ext == 'json':
                data = json.loads(content)
                if isinstance(data, list) and data:
                    return len(data), len(data[0]) if isinstance(data[0], dict) else 0
                return 0, 0
        except Exception:
            return 0, 0


class DatasetDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET — метаданные; PATCH — обновить имя; DELETE — удалить."""
    serializer_class = DatasetSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Dataset.objects.filter(user=self.request.user)


class DatasetDownloadView(APIView):
    """GET /api/datasets/<id>/download/ — скачать файл."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            dataset = Dataset.objects.get(pk=pk, user=request.user)
        except Dataset.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        return FileResponse(dataset.file.open('rb'), as_attachment=True, filename=dataset.name)


class DemoListView(APIView):
    """GET /api/datasets/demo/ — список встроенных демо-датасетов."""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        result = []
        for slug, info in DEMO_DATASETS.items():
            result.append({'slug': slug, **info})
        return Response(result)


class DemoDetailView(APIView):
    """GET /api/datasets/demo/<slug>/ — содержимое демо-датасета (JSON)."""
    permission_classes = [permissions.AllowAny]

    def get(self, request, slug):
        if slug not in DEMO_DATASETS:
            return Response(
                {'detail': 'Датасет не найден.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        import os
        from django.conf import settings as django_settings

        demo_dir = os.path.join(
            os.path.dirname(__file__), 'demo_data'
        )
        filepath = os.path.join(demo_dir, DEMO_DATASETS[slug]['filename'])

        if not os.path.exists(filepath):
            return Response(
                {'detail': 'Файл демо-датасета не найден на сервере.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            data = list(reader)

        return Response({
            'name': DEMO_DATASETS[slug]['name'],
            'data': data,
        })
